from spreadsheet import create_workbook, save_workbook, write_cell_text, write_cell_img
from dumpers import dumpCases, dumpCaseSkins, dumpCasePicture
from openpyxl.styles import PatternFill

list = create_workbook()

row = 10
cols = ["B", "F", "J", "N"]
colspics = ["C", "G", "K", "O"]
colschecks = ["D", "H", "L", "P"]

alp = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

for col in cols:
    list.active.column_dimensions[col].width = 28.67


cases = dumpCases()
cases.reverse()
print(cases)

i = 0
for case in cases:
    skins = dumpCaseSkins(case)
    picture = dumpCasePicture(case)

    write_cell_img(list, colspics[i % 4], str(row - 1), picture)


    for skin in skins:
        write_cell_text(list, cols[i % 4], str(row + 4), skin)
        row += 1

    i += 1
    
    if i != 0 and i % 4 == 0:
        row += 12
    else:
        row -= len(skins)

    
    print(case.split("/")[-1].replace("-", " "))


for row in list.active[f"A1:{alp[list.active.max_column + 1]}{list.active.max_row}"]:
    for cell in row:
        cell.fill = PatternFill("solid", start_color="000000")
        


save_workbook(list, "collection.xlsx")